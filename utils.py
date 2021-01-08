import PyQt5
from PyQt5.QtWidgets import QTableWidgetItem, QWidget
import serial
import threading
from serial.tools import list_ports
from time import sleep
from datetime import datetime
from openpyxl import Workbook
from PyQt5.QtWidgets import QFileDialog

START_STRING = "START"
END_STRING = "END"

styles = {
    "primary_color": 'rgb(2, 117, 216)',
    "danger_color": "rgb(217, 83, 79)",
    "warning_color": "rgb(240, 173, 78)"

}

class SerialConnection():
    def __init__(self, table_widget, status_bar, connection_btn):
        self.connected = False;
        self.available_ports = [];
        self.read_raw_data = '';
        self.read_extracted_data = ''
        self.status_bar = status_bar
        self.data_control = DataControl(table_widget, self.status_bar);
        self.connection_btn = connection_btn;
    
    def connect(self,port,baud):
        self.port = port;
        self.baud = baud;
        self.timeout = 0;
        self.connection = serial.Serial(self.port, self.baud, timeout=0)
        self.connected = True
        self.read_thread = threading.Thread(target=self.read_data_loop,daemon=True)
        self.read_thread.start();

        self.connection_btn.setText("DISCONNECT")
        self.connection_btn.setStyleSheet(
            f"background-color: {styles['danger_color']}")
    
    def disconnect(self):
        self.connection.close();

        self.connection_btn.setText("CONNECT")
        self.connection_btn.setStyleSheet(
            f"background-color: {styles['primary_color']}")
    
    def read_data_loop(self):

        while self.connected:

            chunk_data = self.connection.readline()
            try:
                self.read_raw_data += chunk_data.decode();
            except UnicodeDecodeError:
                print("error")
                
            if b'\n' in chunk_data:
                if (len(self.read_raw_data)):
                    # print(self.read_raw_data);
                    self.read_extracted_data = self.extract_data(
                        self.read_raw_data)
                    if(self.read_extracted_data):

                        # print(self.read_extracted_data)
                        self.data_control.append(self.read_extracted_data);
                    self.read_raw_data = ''
        
                sleep(.01)
            else:
                sleep(.2)

    def check_connection(self):
        if not self.connection.is_open():
            self.connected = False
            self.connection_btn

    def extract_data(self, string_data):
        splited_arr = string_data.split(';')
        # print(splited_arr);
        splited_arr_length = len(splited_arr)
        extracted_data = {}
        if splited_arr[0].find(START_STRING) > -1 and splited_arr[splited_arr_length-1].find(END_STRING) > -1:
            for i in range(1, splited_arr_length-1):
                each_data_string = splited_arr[i]
                each_data_splited_arr = each_data_string.split(':')
                if(len(each_data_splited_arr) == 2):
                    extracted_data[each_data_splited_arr[0]
                                   ] = each_data_splited_arr[1]
        return extracted_data

    def find_ports(self)->list:
        '''
        Used to find all ports connected to the device
        '''
        all_port_tuples = list_ports.comports();
        all_ports = [];
        for ap,_,_ in all_port_tuples:
            all_ports.append(ap);
        self.available_ports = all_ports
        # print(f"from lib: {all_ports}")
        return all_ports
    
    def clear_btn_handle(self):
        self.data_control.clear_all_data();

    def export_exel_btn_handle(self):
        self.data_control.export_to_exel();

    

def show_status(widget,msg):
    widget.showMessage(msg);
    sleep(3);
    widget.showMessage("");


def show_status_thread(widget, msg):
    status_show_thread = threading.Thread(
        target=show_status, args=(widget, msg))
    status_show_thread.start()



def update_port_view(connection,port_selected_box):
    while True:
        ports = connection.find_ports();
        port_selected_box.clear();
        if len(ports):
            print(ports);
            for p in ports:
                port_selected_box.addItem(p);
        sleep(5)

def connection_btn_clicked(connection,btn,portBox,baudBox):
    if not connection.connected:
        port = portBox.currentText();
        baud = baudBox.currentText()
        print(f"port: {port}, baud: {baud}")
        if (len(port)>2):
            connection.connect(port,baud);
        else:
            show_status_thread(connection.status_bar, "No Port Selected!")
    else:
        connection.disconnect();
        connection.connected = False;

    print(f"connected: {connection.connected}");


class DataControl():
    def __init__(self, table_widget, status_bar):
        self.data = [];
        self.table_widget = table_widget;
        self.table_row_count = 0;
        self.status_bar = status_bar;
    
    def append(self, d):
        d['time'] = datetime.now().strftime("%H:%M:%S");
        self.data.append(d);
        self.render_added_data(d);
         # self.render();
        # print(d)
        # print(f"data count: {len(self.data)}")

    def get(self,d):
        return self.data;

    def render_added_data(self,d):
        self.table_widget.insertRow(self.table_row_count);
        self.table_widget.setItem(self.table_row_count,0,QTableWidgetItem(str(d.get("time",""))))
        self.table_widget.setItem(self.table_row_count,1,QTableWidgetItem(str(d.get("current",""))))
        self.table_widget.setItem(self.table_row_count,2,QTableWidgetItem(str(d.get("voltage",""))))
        self.table_widget.setItem(self.table_row_count,3,QTableWidgetItem(str(d.get("power",""))))
        self.table_widget.setItem(self.table_row_count,4,QTableWidgetItem(str(d.get("intensity",""))))
        self.table_widget.setItem(self.table_row_count,5,QTableWidgetItem(str(d.get("Efficiency",""))))
        self.table_widget.setItem(self.table_row_count,6,QTableWidgetItem(str(d.get("solartop",""))))
        self.table_widget.setItem(self.table_row_count,7,QTableWidgetItem(str(d.get("solarBack",""))))
        self.table_widget.setItem(self.table_row_count,8,QTableWidgetItem(str(d.get("collector",""))))
        self.table_widget.setItem(self.table_row_count,9,QTableWidgetItem(str(d.get("waterInlet",""))))
        self.table_widget.setItem(self.table_row_count,10,QTableWidgetItem(str(d.get("wind",""))))
        self.table_row_count +=1
        pass

    def render_all_data(self):
        self.table_widget.setRowCount(0);
        self.table_widget.insertRow(0);
        for row, d in enumerate(self.data):
            self.table_widget.setItem(row,0, QTableWidgetItem(str(d.get('time',""))))
            print(row);
            self.table_widget.insertRow(self.table_widget.rowCount());
        pass

    def clear_all_data(self):
        self.table_widget.setRowCount(0);
        self.data = [];
        self.table_row_count = 0;

    


    def export_to_exel(self):
        print(self.data);
        workbook = Workbook();
        sheet = workbook.active;

        sheet.cell(row=1,column=1).value = "Time";
        sheet.cell(row=1,column=2).value = "Current";
        sheet.cell(row=1,column=3).value = "Voltage";
        sheet.cell(row=1,column=4).value = "Power";
        sheet.cell(row=1,column=5).value = "Intensity";
        sheet.cell(row=1,column=6).value = "Efficiency";
        sheet.cell(row=1,column=7).value = "Solar Top Temp";
        sheet.cell(row=1,column=8).value = "Solar Back Temp";
        sheet.cell(row=1,column=9).value = "Collector Temp";
        sheet.cell(row=1,column=10).value = "Water Inlet Temp";
        sheet.cell(row=1,column=11).value = "Wind Speed";

        for i,d in enumerate(self.data):
            t = 2
            sheet.cell(row=i+t, column=1).value = (d.get("time", ""))
            sheet.cell(row=i+t, column=2).value= (d.get("current", ""))
            sheet.cell(row=i+t, column=3).value = (d.get("voltage", ""))
            sheet.cell(row=i+t, column=4).value = (d.get("power", ""))
            sheet.cell(row=i+t, column=5).value = (d.get("intensity", ""))
            sheet.cell(row=i+t, column=6).value = (d.get("Efficiency", ""))
            sheet.cell(row=i+t, column=7).value = (d.get("solartop", ""))
            sheet.cell(row=i+t, column=8).value = (d.get("solarBack", ""))
            sheet.cell(row=i+t, column=9).value = (d.get("collector", ""))
            sheet.cell(row=i+t, column=10).value = (d.get("waterInlet", ""))
            sheet.cell(row=i+t, column=11).value = (d.get("wind", ""))
        options = QFileDialog.Options();
        widget = QWidget();
        file = QFileDialog.getSaveFileName(widget,"Save file window", "data.xlsx", "Exel Files (*.xlsx)", options=options)
        print(file)
        if file[0] :
            workbook.save(file[0]);
            show_status_thread(self.status_bar,"File saved successfully.");
        else:
            show_status_thread(self.status_bar, "No file location selected!")

